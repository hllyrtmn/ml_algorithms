from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import numpy as np
import pandas as pd

class ExcelReportService:
    
    def __init__(self, filename='ML_Report.xlsx'):
        self.filename = filename
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        
    def create_summary_sheet(self, results):
        ws = self.wb.create_sheet("Özet")
        
        ws['A1'] = 'Makine Öğrenmesi Karşılaştırma Raporu'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f'Oluşturulma Tarihi: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
        
        row = 4
        headers = ['Model', 'Dataset', 'Accuracy', 'Precision', 'Recall', 'F1-Score']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        row = 5
        for result in results:
            ws.cell(row=row, column=1, value=result['model_name'])
            ws.cell(row=row, column=2, value=result['dataset_name'])
            ws.cell(row=row, column=3, value=f"{result['metrics']['accuracy']:.4f}")
            ws.cell(row=row, column=4, value=f"{result['metrics']['precision']:.4f}")
            ws.cell(row=row, column=5, value=f"{result['metrics']['recall']:.4f}")
            ws.cell(row=row, column=6, value=f"{result['metrics']['f1_score']:.4f}")
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        for col in ['C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15
    
    def create_confusion_matrix_sheet(self, model_name, dataset_name, cm, class_names):

        sheet_name = f"{model_name[:10]}_{dataset_name[:10]}"
        ws = self.wb.create_sheet(sheet_name)
        
        ws['A1'] = f'Confusion Matrix: {model_name}'
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f'Dataset: {dataset_name}'
        
        row = 4
        ws.cell(row=row, column=1, value='Gerçek \\ Tahmin')
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        for col, class_name in enumerate(class_names, start=2):
            cell = ws.cell(row=row, column=col, value=class_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for i, class_name in enumerate(class_names):
            row += 1
            cell = ws.cell(row=row, column=1, value=class_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            for j, value in enumerate(cm[i]):
                cell = ws.cell(row=row, column=j+2, value=int(value))
                cell.alignment = Alignment(horizontal="center")
                
                if i == j:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(bold=True)
        
        row += 3
        ws.cell(row=row, column=1, value='Metrik Açıklamaları:')
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        
        explanations = [
            ('Accuracy', 'Doğru tahminlerin toplam tahminlere oranı'),
            ('Precision', 'Pozitif tahminlerin ne kadarının gerçekten pozitif olduğu'),
            ('Recall', 'Gerçek pozitiflerin ne kadarının doğru bulunduğu'),
            ('F1-Score', 'Precision ve Recall\'un ortalaması')
        ]
        
        for i, (metric, explanation) in enumerate(explanations, start=1):
            row += 1
            ws.cell(row=row, column=1, value=f'{metric}:')
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=explanation)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        
        ws.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
    
    def add_algorithm_explanations_sheet(self):
        ws = self.wb.create_sheet("Algoritma Açıklamaları")
        
        ws['A1'] = 'Kullanılan Makine Öğrenmesi Algoritmaları'
        ws['A1'].font = Font(size=14, bold=True)
        
        algorithms = [
            {
                'name': 'Logistic Regression',
                'description': 'İkili veya çok sınıflı sınıflandırma problemleri için kullanılan lineer bir modeldir. Sigmoid fonksiyonu kullanarak olasılık değerleri üretir.',
                'pros': 'Hızlı, yorumlanabilir, az veri ile çalışabilir',
                'cons': 'Lineer ilişkileri varsayar, karmaşık ilişkileri modelleyemeyebilir'
            },
            {
                'name': 'Random Forest',
                'description': 'Birden fazla karar ağacının birleşiminden oluşan ensemble bir öğrenme yöntemidir. Her ağaç rastgele seçilen özellikler üzerinde eğitilir.',
                'pros': 'Overfitting\'e karşı dirençli, özellik önemliliği sağlar, non-lineer ilişkileri yakalar',
                'cons': 'Yorumlanması zor, büyük veri setlerinde yavaş olabilir'
            },
            {
                'name': 'Support Vector Machine (SVM)',
                'description': 'Sınıflar arasında en geniş marjini bulmaya çalışan bir algoritmadır. Kernel trick ile non-lineer problemleri çözebilir.',
                'pros': 'Yüksek boyutlu veride etkili, memory efficient, kernel ile esneklik',
                'cons': 'Büyük veri setlerinde yavaş, hiperparametre seçimi kritik'
            }
        ]
        
        row = 3
        for algo in algorithms:
            ws.cell(row=row, column=1, value=algo['name'])
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            
            row += 1
            ws.cell(row=row, column=1, value='Açıklama:')
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=algo['description'])
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            
            row += 1
            ws.cell(row=row, column=1, value='Avantajlar:')
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=algo['pros'])
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            
            row += 1
            ws.cell(row=row, column=1, value='Dezavantajlar:')
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=algo['cons'])
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            
            row += 2
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 80
    
    def generate_report(self, results):
        self.create_summary_sheet(results)
        
        for result in results:
            self.create_confusion_matrix_sheet(
                result['model_name'],
                result['dataset_name'],
                result['confusion_matrix'],
                result['class_names']
            )
        
        self.add_algorithm_explanations_sheet()
        self.wb.save(self.filename)
        return self.filename